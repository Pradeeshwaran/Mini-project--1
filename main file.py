import openpyxl
import hl
from datetime import datetime

file_name = "patients.xlsx"

try:
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Patient ID", "Name", "Age", "Gender", "Phone",
               "Sleep", "Water", "Exercise", "Heart Rate", "BMI",
               "Score", "Risk", "Date & Time", "Suggestions"])
    wb.save(file_name)


def add_patient():
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    print("\n--- ADD PATIENT ---")
    pid = input("Enter Patient ID: ")

    for row in ws.iter_rows(min_row=2):
        if row[0].value == pid:
            print("ID already exists")
            return

    name = input("Enter Name: ")
    age = int(input("Enter Age: "))
    gender = input("Enter Gender: ")
    phone = input("Enter Phone: ")

    data = {
        "sleep": float(input("Sleep hours: ")),
        "water": float(input("Water intake (l): ")),
        "exercise": int(input("Exercise minutes: ")),
        "heart_rate": int(input("Heart rate: ")),
        "bmi": float(input("BMI: "))
    }

    score = hl.cal_score(data)
    risk = hl.risk_level(score)

    time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    suggestions = hl.suggestions(data)

    ws.append([pid, name, age, gender, phone,
               data["sleep"], data["water"], data["exercise"],
               data["heart_rate"], data["bmi"], score, risk,
               time, suggestions])

    try:
        wb.save(file_name)
        print("✅ Patient saved successfully!")
    except PermissionError:
        print("❌ Close 'patients.xlsx' file and try again!")


def search_patient():
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    pid = input("\nEnter Patient ID to search: ")

    for row in ws.iter_rows(min_row=2):
        if row[0].value == pid:
            print("\n--- PATIENT REPORT ---")
            print("ID:", row[0].value)
            print("Name:", row[1].value)
            print("Age:", row[2].value)
            print("Gender:", row[3].value)
            print("Phone:", row[4].value)
            print("Sleep:", row[5].value)
            print("Water:", row[6].value)
            print("Exercise:", row[7].value)
            print("Heart Rate:", row[8].value)
            print("BMI:", row[9].value)
            print("Score:", row[10].value)
            print("Risk:", row[11].value)
            print("Date & Time:", row[12].value)
            print("Suggestions:", row[13].value)
            return

    print("❌ Patient not found!")


def show_all():
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    print("\n--- ALL PATIENTS ---")
    for row in ws.iter_rows(min_row=2):
        print("\n-----------------------------")
        print("ID:", row[0].value)
        print("Name:", row[1].value)
        print("Age:", row[2].value)
        print("Gender:", row[3].value)
        print("Phone:", row[4].value)
        print("Sleep:", row[5].value)
        print("Water:", row[6].value)
        print("Exercise:", row[7].value)
        print("Heart Rate:", row[8].value)
        print("BMI:", row[9].value)
        print("Score:", row[10].value)
        print("Risk:", row[11].value)
        print("Date & Time:", row[12].value)
        print("Suggestions:", row[13].value)

while True:
    print("\n===== DIGITAL HEALTH MONITOR & RISK PREDICTOR  =====")
    print("1. Add Patient")
    print("2. Search Patient by ID")
    print("3. Show All Patients")
    print("4. Exit")

    try:
        choice = int(input("Enter choice: "))
    except:
        print("Invalid input!")
        continue

    if choice == 1:
        add_patient()
    elif choice == 2:
        search_patient()
    elif choice == 3:
        show_all()
    elif choice == 4:
        print("Exiting...")
        break
    else:
        print("Invalid choice!")
